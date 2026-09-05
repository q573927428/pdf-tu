"""PDF discovery, rendering and catalog output."""
from __future__ import annotations

import csv, hashlib, logging, os, re, time, shutil, json, urllib.request, urllib.parse, urllib.error, base64, mimetypes, threading, subprocess, sys
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pymupdf
import yaml
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

LOG = logging.getLogger(__name__)
HEADERS = ["序号", "年级", "学期", "科目", "分类", "PDF 文件名称", "生成文案", "封面图链接", *[f"图 {i}" for i in range(1, 6)], "操作"]

@dataclass
class Settings:
    source_root: Path; output_root: Path; grade: str; semester: str
    watermark: dict[str, Any]; render: dict[str, Any]; processing: dict[str, Any]; table: dict[str, Any]; ai: dict[str, Any] = field(default_factory=dict)

def load_settings(path: str | Path) -> Settings:
    p = Path(path).resolve(); raw = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    def req(name):
        if not raw.get(name): raise ValueError(f"缺少配置项: {name}")
        return raw[name]
    source = Path(req("source_root")).expanduser()
    if not source.is_absolute(): source = (p.parent / source).resolve()
    out = Path(req("output_root")).expanduser()
    if not out.is_absolute(): out = (p.parent / out).resolve()
    render = {"dpi": 150, "max_pages": 5, "max_width": None, "max_height": None, "output_width": None, "png_optimize": True, **raw.get("render", {})}
    if not (1 <= int(render["dpi"]) <= 1200): raise ValueError("render.dpi 必须在 1~1200")
    if not (1 <= int(render["max_pages"]) <= 100): raise ValueError("render.max_pages 必须在 1~100")
    for key in ("max_width", "max_height", "output_width"):
        if render.get(key) is not None and int(render[key]) < 1: raise ValueError(f"render.{key} 必须为正整数或 null")
    wm = {"enabled": True, "text": "幼升小", "opacity": 90, "position": "bottom_right", "font_size": 28, **raw.get("watermark", {})}
    if wm["position"] not in {"bottom_right", "bottom_left", "center"}: raise ValueError("watermark.position 不支持")
    wm["opacity"] = max(0, min(255, int(wm["opacity"])))
    table = {"xlsx": "pdf_catalog.xlsx", "csv": "pdf_catalog.csv", "include_metadata_title": False, **raw.get("table", {})}
    ai = {"enabled": False, "endpoint": "https://ark.cn-beijing.volces.com/api/v3", "api_key": "", "model": "", "copy_model": "", "image_model": "", "reference_image": "", "reference_images": [], "generate_copy": False, "generate_cover": False, "timeout": 60, **raw.get("ai", {})}
    reference_image = str(ai.get("reference_image") or "").strip()
    if reference_image and not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", reference_image):
        ref_path = Path(reference_image).expanduser()
        if not ref_path.is_absolute():
            ai["reference_image"] = str((p.parent / ref_path).resolve())
    reference_images = ai.get("reference_images") or []
    if isinstance(reference_images, str):
        reference_images = [reference_images]
    resolved_images = []
    for item in reference_images:
        item = str(item).strip()
        if item and not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://|^data:", item):
            image_path = Path(item).expanduser()
            if not image_path.is_absolute():
                item = str((p.parent / image_path).resolve())
        if item:
            resolved_images.append(item)
    ai["reference_images"] = resolved_images
    return Settings(source, out, str(req("grade")), str(req("semester")), wm, render, {"max_pdfs": None, **raw.get("processing", {})}, table, ai)

def discover(root: Path) -> list[Path]:
    return sorted((p for p in root.rglob("*") if p.is_file() and p.suffix.lower() == ".pdf"), key=lambda p: str(p.relative_to(root)).lower())

def safe_name(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._") or "pdf"
    return value[:100]

def safe_dir_name(value: str) -> str:
    """保留中文目录名，仅替换 Windows 不允许的字符。"""
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .")
    return value[:100] or "未分类"

def directory_fields(pdf: Path, settings: Settings) -> tuple[str, str, str]:
    """从 PDF 相对路径提取学期、科目和分类目录。"""
    parts = list(pdf.relative_to(settings.source_root).parts[:-1])
    semester = settings.semester
    if parts and parts[0] in {"上册", "下册"}:
        semester, parts = parts[0], parts[1:]
    return semester, (parts[0] if parts else ""), (parts[1] if len(parts) > 1 else "")

def _font(size: int):
    for p in (r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\simhei.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"):
        if Path(p).exists():
            try: return ImageFont.truetype(p, size)
            except OSError: pass
    return ImageFont.load_default()

def _fingerprint(pdf: Path, settings: Settings) -> str:
    st = pdf.stat(); payload = f"{st.st_size}:{st.st_mtime_ns}:{settings.render['dpi']}:{settings.render['max_pages']}:{settings.watermark}"; return hashlib.sha1(payload.encode()).hexdigest()[:12]

def process_pdf(pdf: Path, settings: Settings, sequence: int, watermark=True) -> tuple[str, list[str], str | None]:
    rel = pdf.relative_to(settings.source_root); semester, subject, category = directory_fields(pdf, settings)
    token = hashlib.sha1(str(rel).encode("utf-8")).hexdigest()[:8]
    folder = settings.output_root / "images" / safe_dir_name(settings.grade) / safe_dir_name(semester) / safe_dir_name(subject or "未分类") / safe_dir_name(category or "未分类") / f"{sequence:06d}"; folder.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(pdf, folder / pdf.name)
        doc = pymupdf.open(pdf); title = (doc.metadata.get("title") or "").strip() if settings.table.get("include_metadata_title") else ""
        title = title or pdf.stem; count = min(len(doc), int(settings.render["max_pages"])); paths=[]; fp=_fingerprint(pdf, settings)
        marker = folder / ".fingerprint"
        for i in range(count):
            out = folder / f"page-{i+1:02d}.png"
            if out.exists() and marker.exists() and marker.read_text() == fp: paths.append(str(out.relative_to(settings.output_root)).replace(os.sep, "/")); continue
            pix = doc.load_page(i).get_pixmap(matrix=pymupdf.Matrix(float(settings.render["dpi"])/72, float(settings.render["dpi"])/72), alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            if watermark and settings.watermark.get("enabled") and settings.watermark.get("text"):
                layer=Image.new("RGBA", img.size, (0,0,0,0)); d=ImageDraw.Draw(layer); f=_font(int(settings.watermark["font_size"])); box=d.textbbox((0,0), settings.watermark["text"], font=f); w,h=box[2]-box[0],box[3]-box[1]; margin=max(1,int(min(img.size)*.03)); pos={"bottom_right":(img.width-w-margin,img.height-h-margin),"bottom_left":(margin,img.height-h-margin),"center":((img.width-w)//2,(img.height-h)//2)}[settings.watermark["position"]]; d.text(pos, settings.watermark["text"], font=f, fill=(255,0,0,int(settings.watermark["opacity"]))) ; img=Image.alpha_composite(img.convert("RGBA"),layer).convert("RGB")
            # output_width 为固定宽度并按比例自动计算高度；否则按 max_width/max_height 等比限制。
            ow, mw, mh = settings.render.get("output_width"), settings.render.get("max_width"), settings.render.get("max_height")
            target = None
            if ow:
                target = (int(ow), max(1, round(img.height * int(ow) / img.width)))
            else:
                scale = min((int(mw) / img.width) if mw else 1.0, (int(mh) / img.height) if mh else 1.0, 1.0)
                if scale < 1:
                    target = (max(1, round(img.width * scale)), max(1, round(img.height * scale)))
            if target: img = img.resize(target, Image.Resampling.LANCZOS)
            tmp=out.with_suffix(".tmp.png"); img.save(tmp, "PNG", optimize=bool(settings.render.get("png_optimize", True))); os.replace(tmp,out); paths.append(str(out.relative_to(settings.output_root)).replace(os.sep,"/"))
        marker.write_text(fp, encoding="ascii"); doc.close(); return title, paths, None
    except Exception as exc:
        return pdf.stem, [], f"{type(exc).__name__}: {exc}"

def _output_folder(pdf: Path, settings: Settings, sequence: int) -> Path:
    """返回 PDF 页面图片所在目录，封面图与页面图共用该目录。"""
    rel = pdf.relative_to(settings.source_root)
    semester, subject, category = directory_fields(pdf, settings)
    return (settings.output_root / "images" / safe_dir_name(settings.grade) /
            safe_dir_name(semester) / safe_dir_name(subject or "未分类") /
            safe_dir_name(category or "未分类") / f"{sequence:06d}")

def _download_cover(url: str, pdf: Path, settings: Settings, sequence: int) -> str:
    """下载封面到页面图片目录，并返回相对 output_root 的路径。"""
    if not url:
        return ""
    folder = _output_folder(pdf, settings, sequence)
    folder.mkdir(parents=True, exist_ok=True)
    # 每个 PDF 的封面文件名固定，避免文件名变化影响后续引用。
    out = folder / "cover.png"
    if url.startswith("data:"):
        header, encoded = url.split(",", 1)
        mime = header[5:].split(";", 1)[0]
        content = base64.b64decode(encoded)
    else:
        req = urllib.request.Request(url, headers={"User-Agent": "pdf-catalog"})
        with urllib.request.urlopen(req, timeout=float(settings.ai.get("timeout", 60))) as res:
            content = res.read()
    # 统一转 PNG，确保固定扩展名和 Excel 嵌入兼容性。
    tmp = folder / "cover.png.tmp"
    try:
        tmp.write_bytes(content)
        with Image.open(tmp) as image:
            image.convert("RGB").save(tmp.with_suffix(".converted.png"), "PNG")
        os.replace(tmp.with_suffix(".converted.png"), out)
    finally:
        if tmp.exists():
            tmp.unlink()
    return str(out.relative_to(settings.output_root)).replace(os.sep, "/")

def _image_input(value: str, settings: Settings) -> str:
    """将本地图片转换为 data URI；远程 URL 保持不变。"""
    value = str(value or "").strip()
    if not value or value.startswith("data:") or re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", value):
        return value
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = settings.output_root / path
    if not path.is_file():
        raise FileNotFoundError(f"参考图片不存在: {path}")
    mime = mimetypes.guess_type(path.name)[0] or "image/png"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _doubao(settings: Settings, messages: list[dict[str, str]], *, image=False, image_references: list[str] | None = None) -> Any:
    """调用豆包 Ark 的 OpenAI 兼容接口；未配置时明确报错。"""
    model_key = "image_model" if image else "copy_model"
    model = settings.ai.get(model_key) or settings.ai.get("model")
    if not settings.ai.get("enabled") or not settings.ai.get("api_key") or not model:
        raise RuntimeError(f"未配置 ai.enabled、ai.api_key 或 ai.{model_key}（也可使用兼容字段 ai.model）")
    if image:
        endpoint = str(settings.ai.get("endpoint", "")).rstrip("/") + "/images/generations"
        payload = {"model": model, "prompt": messages[-1].get("content", ""), "size": settings.ai.get("image_size", "1024x1536"), "n": 1}
        configured = settings.ai.get("reference_images") or []
        if isinstance(configured, str):
            configured = [configured]
        refs = [str(x).strip() for x in configured if str(x).strip()]
        single = str(settings.ai.get("reference_image") or "").strip()
        if single:
            refs.insert(0, single)
        refs.extend(str(x).strip() for x in (image_references or []) if str(x).strip())
        # 保持顺序并去重；为控制请求体大小，封面最多携带 3 张参考图。
        # 该数量包含配置中的固定参考图和当前 PDF 页面图。
        refs = list(dict.fromkeys(refs))[:3]
        if refs:
            encoded_refs = [_image_input(ref, settings) for ref in refs]
            # 只有一张时保持单图兼容格式；PDF 页面存在时自然使用多图数组。
            payload["image"] = encoded_refs if len(encoded_refs) > 1 else encoded_refs[0]
    else:
        endpoint = str(settings.ai.get("endpoint", "")).rstrip("/") + "/chat/completions"
        payload = {"model": model, "messages": messages}
    req = urllib.request.Request(endpoint, data=json.dumps(payload, ensure_ascii=False).encode(), headers={"Authorization": "Bearer " + settings.ai["api_key"], "Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=float(settings.ai.get("timeout", 60))) as res:
        return json.loads(res.read().decode("utf-8"))


# 与业务侧提示词保持一致，避免 AI 生成时因提示词过度压缩而遗漏审核红线。
COPY_RED_LINES = """内容红线——以下一律禁止出现：
1. 绝对化/夸大用语：最、第一、唯一、顶级、全网独家、绝无仅有、100%、满分、包过、保分、稳上、必拿A、成绩暴涨、速成、永久免费等；
2. 承诺提分/保证效果：不得写“稳拿A”“提X分”“保证进步”，改为“帮助查漏补缺”“巩固基础”“考前更安心”；
3. 过度贩卖焦虑：不得写“别人家孩子都会了，你家还在玩”“再不练就完了”“一步落后步步落后”等恐吓式表述；紧迫感只能用真实时间节点正向表达；
4. 虚假稀缺/诱导：不得虚构“马上删”“最后几份”“不转就没了”“免费送”等不实信息；
5. 站外引流：不得出现加微信、私信领、点链接、群号等导流行为；
6. 不实头衔：不得使用“官方”“内部”“绝密”“独家整理”等无法证实的表述；
7. 拉踩贬低：不得贬低其他资料、机构或老师；
8. 版权风险：不得出现盗版、破解、翻印等字眼，统一称“学习资料/练习题”。"""

COPY_FORBIDDEN_TERMS = (
    "最", "第一", "唯一", "顶级", "全网独家", "绝无仅有", "100%", "满分", "包过", "保分",
    "稳上", "必拿A", "成绩暴涨", "速成", "永久免费", "稳拿A", "提分", "保证进步", "马上删",
    "最后几份", "不转就没了", "免费送", "加微信", "私信领", "点链接", "群号", "官方", "内部",
    "绝密", "独家整理", "盗版", "破解", "翻印", "别人家孩子都会了，你家还在玩", "再不练就完了",
    "一步落后步步落后",
)


def _copy_prompt(filename: str, category: str, grade_subject: str) -> str:
    return f"""【角色】你是一位熟悉《广告法》及抖音、小红书内容审核规范的教辅资料营销文案专家。文案必须合规、真实、正向，同时保留适度紧迫感和行动感，让宝妈想保存、想打印。
【任务】根据【文件名】【资料类型】【年级科目】，生成一个不超过20个字的标题，再生成一条正文不超过150字（含标点）的营销文案，并在正文末尾添加1至5个相关话题；话题不计入150字限制。
【输入】文件名：{filename}；资料类型：{category or '学习资料'}；年级科目：{grade_subject}
{COPY_RED_LINES}
【写作要求】口语化、有画面感，像懂行的学姐/老师提醒宝妈；紧迫感使用真实时间节点+轻行动建议，不靠恐吓；开头点明场景/痛点，中间说明资料覆盖内容和帮助，结尾使用一个轻行动指令（打印、保存、每天练一页）；必须紧扣文件名具体内容，避免通用套话；可用感叹号、省略号，emoji不超过1个；话题须使用#开头，彼此用空格分隔，放在正文最后。
【标题要求】标题必须紧扣资料内容，简洁有吸引力，不得超过20个字（含标点），不要带话题标签。
【输出格式】严格只输出两行，不加解释、不加引号：
标题：<20个字以内的标题>
正文：<正文和末尾话题>"""


def _copy_result_parts(result: str, filename: str) -> tuple[str, str]:
    """解析模型返回的标题和正文，并兼容偶发的单行输出。"""
    cleaned = str(result or "").strip().strip("`\"'“”")
    title_match = re.search(r"(?:^|\n)\s*标题\s*[：:]\s*(.+?)(?=\n|\s+正文\s*[：:]|$)", cleaned)
    body_match = re.search(r"(?:^|\n|\s)正文\s*[：:]\s*(.+)$", cleaned, re.S)
    if not title_match and not body_match and "\n" in cleaned:
        title, body = (part.strip() for part in cleaned.split("\n", 1))
    else:
        title = title_match.group(1).strip() if title_match else ""
        body = body_match.group(1).strip() if body_match else re.sub(r"^\s*标题\s*[：:]\s*[^\n]+\n?", "", cleaned).strip()
    body = re.sub(r"\s*\n\s*", "", body)
    for term in COPY_FORBIDDEN_TERMS:
        title = title.replace(term, "")
        body = body.replace(term, "")
    title = re.sub(r"^[#\s]+|[#\s]+$", "", title).strip("`\"'“”")
    if not title:
        candidate = re.split(r"[!！?？。，,\s#]", body, maxsplit=1)[0].strip()
        title = candidate or Path(filename).stem
    return title[:20].rstrip(), body


def _copy_topics(result: str, category: str, grade_subject: str) -> tuple[str, list[str]]:
    """Extract trailing hashtags so they do not count toward the copy limit."""
    match = re.search(r"(?:^|\s)((?:#[^\s#]+(?:\s+|$)){1,})$", result)
    if match:
        topics = re.findall(r"#[^\s#]+", match.group(1))[:5]
        body = result[:match.start(1)].rstrip(" ，。；;、")
        if topics:
            return body, topics

    topics = []
    for value in (category, grade_subject, "学习资料", "每日练习"):
        topic = re.sub(r"\s+", "", str(value or "")).lstrip("#")
        if topic and topic not in {item[1:] for item in topics}:
            topics.append(f"#{topic}")
        if len(topics) == 5:
            break
    return result, topics[:5]

def generate_copy(settings: Settings, filename: str, category: str, grade_subject: str) -> str:
    prompt = _copy_prompt(filename, category, grade_subject)
    data = _doubao(settings, [{"role": "user", "content": prompt}])
    content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
    if isinstance(content, list): content = "".join(x.get("text", "") for x in content if isinstance(x, dict))
    # 防止模型偶尔带回 Markdown 包裹、超出字数要求或重复出现明确禁用词。
    title, result = _copy_result_parts(str(content), filename)
    body, topics = _copy_topics(result, category, grade_subject)
    copy = f"{body[:150].rstrip()} {' '.join(topics)}".strip()
    return f"{title}\n{copy}"

def generate_cover(settings: Settings, copy: str, name: str, category: str, grade_subject: str, page_images: list[str] | None = None) -> str:
    prompt = f"""【角色】你是一位儿童教辅类学习资料封面设计师，擅长为宝妈群体设计温暖、干净、有手账感的竖版封面。
【任务】根据封面文案、资料名称、亮点、年级科目生成一张竖版封面图。
【固定段——背景与底部每张封面必须严格照抄，一个字都不能改】
竖版儿童学习资料分享封面，日系治愈手账风，奶油白与浅燕麦色纸张底纹铺满背景，轻微纸纹颗粒与手绘笔触质感，雾蓝、暖粉、鹅黄三色低饱和点缀，暖色柔和光线与轻微纸张阴影，画面无水印无平台标识。
- 底部约10%提示区：一行灰棕色手写小字（如“打印版·趁早存好”）
- 左下角头像保护区（最高优先级）：参考图中左下角的圆形头像/徽章是不可修改的固定素材，必须原样保留。
【顶部标题区——位置永远固定，样式与文字随文案可变】
- 位置固定：标题始终位于画面顶部约15%高度的居中区域，标题下方保留一条鹅黄色手绘波浪下划线
- 文字可变：标题内容从文案中提取最抓眼的一句（≤10字），随不同文案变化
- 样式可变：根据标题字数与文案语气，从以下样式中选用或微调：a. 荧光笔高亮；b. 圆角标签+印章；c. 描边贴纸字；d. 便签标题条。标题多时可缩小或折行，但位置与下划线不动。
【可变段——中部内容区，自由发挥，不固定模板】
请根据资料名称、亮点短句和标题语气，自主设计富有变化的手账式版面。可灵活组合或改造便签、单张卡片、错落拼贴、文件夹、练习纸、书本、铅笔、星星、纸胶带等元素；每张封面可改变卡片数量、大小、位置、倾斜角度、留白比例和装饰组合，避免重复使用相同的版式框架。资料名称与亮点要有清晰主次、方便阅读；年级科目统一放在中下部的米黄色横条标签卡中，位置可有小幅变化但需保持稳定识别。
【本张可变内容】封面文案：“{copy}”；资料名称：“{name}”；亮点短句：“{category or '查漏补缺·每日一练'}”；年级科目：“{grade_subject}”。
【规则】所有待渲染中文用“”标出并写清位置；全图不超过4处主要文字块，每处≤10字；文字清晰可读、笔画完整；严禁水印、平台logo。背景质感、主色调、字体气质和信息层级保持统一，版面结构、卡片形态、装饰元素与留白方式可自然变化；画面要像同一套系列资料，但不要每张都长得一样。"""
    data = _doubao(settings, [{"role": "user", "content": prompt}], image=True, image_references=(page_images or []))
    def find(v):
        if isinstance(v, str) and (v.startswith("http://") or v.startswith("https://") or v.startswith("data:image/")): return v
        if isinstance(v, dict):
            # 兼容部分图片模型返回的 base64 字段。
            encoded = v.get("b64_json")
            if isinstance(encoded, str) and encoded:
                return "data:image/png;base64," + encoded
            for x in v.values():
                got = find(x)
                if got: return got
        if isinstance(v, list):
            for x in v:
                got = find(x)
                if got: return got
        return ""
    return find(data)

def _status_path(settings: Settings) -> Path:
    return settings.output_root / ".publish_status.json"

def _load_publish_status(settings: Settings) -> dict[str, str]:
    path = _status_path(settings)
    if not path.is_file(): return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return {str(k): ("已发布" if v == "已发布" else "未发布") for k, v in data.items()}
    except (OSError, ValueError, TypeError):
        LOG.warning("发布状态文件读取失败，将使用未发布状态: %s", path)
        return {}

def _save_publish_status(settings: Settings, statuses: dict[str, str]) -> None:
    path = _status_path(settings); tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(statuses, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, path)

def _update_status_in_html(settings: Settings, sequence: int, status: str) -> None:
    """只更新 HTML 中指定序号的状态按钮，避免切换状态时重建大型 XLSX。"""
    html_name = settings.table.get("html", "pdf_catalog.html")
    html_path = settings.output_root / html_name
    if not html_path.is_file():
        return
    content = html_path.read_text(encoding="utf-8")
    sequence_text = re.escape(str(sequence))
    status_text = re.escape(status)
    css_class = "published" if status == "已发布" else "unpublished"
    pattern = re.compile(
        rf'(<button class="status-btn )(?:published|unpublished)('
        rf'" data-sequence="{sequence_text}" data-status=")(?:已发布|未发布)('
        rf'" onclick="toggleStatus\(this\)">)(?:已发布|未发布)(</button>)'
    )
    replacement = rf'\g<1>{css_class}\g<2>{status_text}\g<3>{status}\g<4>'
    updated, count = pattern.subn(replacement, content, count=1)
    if count:
        html_path.write_text(updated, encoding="utf-8")
    else:
        LOG.warning("HTML 中未找到序号 %s 的发布状态按钮", sequence)

def write_tables(rows: list[dict[str, Any]], settings: Settings, errors: list[dict[str,str]], elapsed: float, details: list[str] | None = None) -> None:
    settings.output_root.mkdir(parents=True, exist_ok=True); xlsx=settings.output_root/settings.table["xlsx"]; csvp=settings.output_root/settings.table["csv"]
    wb=Workbook(); ws=wb.active; ws.title="PDF目录"; ws.append(HEADERS); ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}{len(rows)+1}"
    for c in ws[1]: c.font=Font(bold=True); c.alignment=Alignment(horizontal="center")
    for row in rows: ws.append([row.get(h,"") for h in HEADERS])
    # XLSX 中所有图片列均保留路径文字，不嵌入图片；缩略图仅用于 HTML 展示。
    # 按字段设置列宽，避免“生成文案”占用过多横向空间。
    widths = {"序号": 8, "年级": 10, "学期": 10, "科目": 10, "分类": 12,
              "PDF 文件名称": 26, "生成文案": 32, "封面图链接": 20}
    for h in HEADERS:
        widths.setdefault(h, 18)
    for col, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths[h]
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment=Alignment(vertical="top", wrap_text=True)
    try:
        wb.save(xlsx)
        with csvp.open("w", newline="", encoding="utf-8-sig") as f: w=csv.DictWriter(f, fieldnames=HEADERS); w.writeheader(); w.writerows(rows)
        with (settings.output_root/"errors.csv").open("w", newline="", encoding="utf-8-sig") as f: w=csv.DictWriter(f, fieldnames=["path","stage","error"]); w.writeheader(); w.writerows(errors)
        # 生成单页 HTML 表格，便于浏览器查看、搜索和打印。
        html_name = settings.table.get("html", "pdf_catalog.html")
        htmlp = settings.output_root / html_name
        def esc(value: Any) -> str:
            import html
            return html.escape(str(value or ""))
        def copy_control(value: str, sequence: str = "", refresh: bool = False, folder: bool = False, titled: bool = False) -> str:
            title, copy = (value.split("\n", 1) if titled and "\n" in value else ("", value))
            control = ""
            if title:
                control += (f'<div class="copy-title-row"><strong class="copy-title">{esc(title)}</strong>'
                            f'<button class="copy-btn title-copy-btn" type="button" data-copy="{esc(title)}" '
                            f'onclick="copyText(this)" title="复制标题" aria-label="复制标题">⧉</button></div>')
            control += (f'<div class="copy-body"><span class="copy-value">{esc(copy)}</span>'
                        f'<button class="copy-btn" type="button" data-copy="{esc(copy)}" '
                        f'onclick="copyText(this)" title="复制文案" aria-label="复制文案">⧉</button>')
            if folder:
                control += (f'<button class="copy-btn folder-btn" type="button" data-sequence="{esc(sequence)}" '
                            f'onclick="openFolder(this)" title="打开生成文件目录" aria-label="打开生成文件目录">📁</button>')
            if refresh:
                control += (f'<button class="copy-btn" type="button" data-action="copy" '
                            f'data-sequence="{esc(sequence)}" onclick="generate(this)" '
                            f'title="重新生成文案" aria-label="重新生成文案">↻</button>')
            return control + "</div>"
        head = "".join(f"<th>{esc(h)}</th>" for h in HEADERS)
        body = []
        statuses = _load_publish_status(settings)
        for row in rows:
            cells = []
            for h in HEADERS:
                value = str(row.get(h, "") or "")
                sequence = esc(row.get("序号", ""))
                if h == "封面图链接" and value:
                    cells.append(f'<td><a href="{esc(value)}" target="_blank"><img src="{esc(value)}" alt="封面" loading="lazy"></a>'
                                 f'<br><button class="copy-btn" type="button" data-action="cover" data-sequence="{sequence}" '
                                 f'onclick="generate(this)" title="重新生成封面" aria-label="重新生成封面">↻</button></td>')
                elif h == "封面图链接":
                    cells.append(f'<td><button class="generate-btn" data-action="cover" data-sequence="{sequence}" onclick="generate(this)">生成封面</button></td>')
                elif h == "生成文案" and not value:
                    cells.append(f'<td><button class="generate-btn" data-action="copy" data-sequence="{sequence}" onclick="generate(this)">生成文案</button></td>')
                elif h in {"PDF 文件名称", "生成文案"}:
                    cells.append(f"<td>{copy_control(value, sequence, h == '生成文案', h == 'PDF 文件名称', h == '生成文案')}</td>")
                elif h == "操作":
                    status = statuses.get(str(row.get("序号", "")), "未发布")
                    cells.append(f'<td><button class="status-btn {"published" if status == "已发布" else "unpublished"}" data-sequence="{sequence}" data-status="{status}" onclick="toggleStatus(this)">{status}</button></td>')
                elif h.startswith("图") and value:
                    # 单页表格中直接显示缩略图，不展示图片路径链接；点击缩略图可打开原图。
                    cells.append(f'<td><a href="{esc(value)}" target="_blank"><img src="{esc(value)}" alt="{esc(h)}" loading="lazy"></a></td>')
                else:
                    cells.append(f"<td>{esc(value)}</td>")
            # 将序号写入行本身，异步生成请求完成后可重新定位当前 DOM，
            # 避免并发请求中某个请求替换单元格导致旧 button 引用失效。
            body.append(f'<tr data-sequence="{esc(row.get("序号", ""))}">' + "".join(cells) + "</tr>")
        htmlp.write_text("""<!doctype html>
<html lang=\"zh-CN\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">
<title>PDF 目录</title><style>
body{font-family:Arial,\"Microsoft YaHei\",sans-serif;margin:20px;color:#222}h1{font-size:22px}#search{width: min(520px,100%);padding:9px 12px;border:1px solid #bbb;border-radius:6px;margin:0 0 14px;font-size:14px}
.table-wrap{overflow:auto;border:1px solid #ddd}table{border-collapse:collapse;width:100%;min-width:1500px;font-size:13px}th,td{border:1px solid #ddd;padding:7px;vertical-align:top;line-height:1.4}th{position:sticky;top:0;background:#f3f5f7;white-space:nowrap}td img{max-width:90px;max-height:120px;margin-top:4px}a{color:#1769aa;word-break:break-all}th:nth-child(7),td:nth-child(7){width:360px;min-width:180px;max-width:368px;white-space:normal;overflow-wrap:anywhere}.generate-btn,.copy-btn,.page-btn,.status-btn{padding:6px 10px;border:1px solid #1769aa;border-radius:5px;background:#fff;color:#1769aa;cursor:pointer;white-space:nowrap}.generate-btn:disabled,.page-btn:disabled,.status-btn:disabled{opacity:.55;cursor:wait}.copy-btn{padding:1px 5px;margin-top:2px;font-size:14px;line-height:1.1;vertical-align:middle;border:0}.copy-btn:hover,.page-btn:not(:disabled):hover{background:#eaf4ff}.copy-title-row{display:flex;align-items:center;gap:4px;margin-bottom:6px}.copy-title{font-size:14px;color:#111}.title-copy-btn{flex:0 0 auto;margin-top:0}.copy-body{color:#444}.status-btn.published{border-color:#299447;color:#207a39;background:#effaf1}.status-btn.unpublished{border-color:#999;color:#666;background:#fafafa}.notice{color:#777;font-size:12px;margin:-6px 0 14px}.pagination{display:flex;align-items:center;justify-content:center;gap:12px;padding:14px}.page-info{color:#666;font-size:13px}.toast{position:fixed;left:50%;top:24px;transform:translateX(-50%);background:#333;color:#fff;padding:9px 16px;border-radius:5px;z-index:10;opacity:0;transition:opacity .2s}.toast.show{opacity:1}
th:nth-child(6),td:nth-child(6){width:260px;min-width:0;max-width:260px;white-space:normal;overflow-wrap:anywhere}
@media print{#search{display:none}.table-wrap{overflow:visible;border:0}table{min-width:0;font-size:8px}th{position:static}td img{max-width:45px;max-height:60px}}
</style></head><body><h1>PDF 目录（分页表格）</h1><p class=\"notice\">每页显示 50 条。缺少封面图或文案时，可点击对应按钮生成。首次使用请在项目目录运行：<code>pdf-catalog serve --config config.yaml</code></p><div id=\"toast\" class=\"toast\" role=\"status\"></div><input id=\"search\" placeholder=\"输入关键词筛选…\" oninput=\"filterRows()\"><div class=\"table-wrap\"><table><thead><tr>""" + head + """</tr></thead><tbody id=\"rows\">""" + "".join(body) + """</tbody></table></div><div class=\"pagination\"><button id=\"prev-page\" class=\"page-btn\" type=\"button\" onclick=\"changePage(-1)\">上一页</button><span id=\"page-info\" class=\"page-info\"></span><button id=\"next-page\" class=\"page-btn\" type=\"button\" onclick=\"changePage(1)\">下一页</button></div><script>
const API_BASE='http://127.0.0.1:8765';
let currentPage=1;const pageSize=50;
function filteredRows(){const q=document.getElementById('search').value.toLowerCase();return [...document.querySelectorAll('#rows tr')].filter(r=>r.innerText.toLowerCase().includes(q))}
function renderPage(){const matched=filteredRows(), totalPages=Math.max(1,Math.ceil(matched.length/pageSize));currentPage=Math.min(currentPage,totalPages);const start=(currentPage-1)*pageSize;const selected=new Set(matched.slice(start,start+pageSize));document.querySelectorAll('#rows tr').forEach(r=>r.style.display=selected.has(r)?'':'none');document.getElementById('page-info').textContent=`第 ${currentPage} / ${totalPages} 页（共 ${matched.length} 条）`;document.getElementById('prev-page').disabled=currentPage<=1;document.getElementById('next-page').disabled=currentPage>=totalPages}
function changePage(delta){currentPage+=delta;renderPage()}
function filterRows(){currentPage=1;renderPage()}
async function toggleStatus(button){const oldStatus=button.dataset.status;const status=oldStatus==='已发布'?'未发布':'已发布';button.disabled=true;try{const response=await fetch(API_BASE+'/api/status',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({sequence:button.dataset.sequence,status})});const result=await response.json();if(!response.ok||!result.ok)throw new Error(result.error||'状态保存失败');button.dataset.status=status;button.textContent=status;button.classList.toggle('published',status==='已发布');button.classList.toggle('unpublished',status==='未发布');showToast(status)}catch(error){showToast(error.message)}finally{button.disabled=false}}
function showToast(message){const toast=document.getElementById('toast');toast.textContent=message;toast.classList.add('show');clearTimeout(window.toastTimer);window.toastTimer=setTimeout(()=>toast.classList.remove('show'),1800)}
async function copyText(button){const value=button.dataset.copy||'';try{if(navigator.clipboard&&window.isSecureContext){await navigator.clipboard.writeText(value)}else{const area=document.createElement('textarea');area.value=value;area.style.position='fixed';area.style.opacity='0';document.body.appendChild(area);area.focus();area.select();document.execCommand('copy');area.remove()}showToast('复制成功')}catch(error){showToast('复制失败，请手动复制')}}
async function openFolder(button){const sequence=button.dataset.sequence||'';button.disabled=true;try{const response=await fetch(API_BASE+'/api/open-folder',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({sequence})});const result=await response.json();if(!response.ok||!result.ok)throw new Error(result.error||'打开目录失败');showToast('已打开生成文件目录')}catch(error){showToast(error.message)}finally{button.disabled=false}}
function setCopyCell(cell,value,sequence){const split=value.indexOf('\\n'),title=split>=0?value.slice(0,split):'',copy=split>=0?value.slice(split+1):value;cell.innerHTML=(title?'<div class="copy-title-row"><strong class="copy-title"></strong><button class="copy-btn title-copy-btn" type="button" title="复制标题" aria-label="复制标题" onclick="copyText(this)">⧉</button></div>':'')+'<div class="copy-body"><span class="copy-value"></span><button class="copy-btn body-copy-btn" type="button" title="复制文案" aria-label="复制文案" onclick="copyText(this)">⧉</button><button class="copy-btn" type="button" data-action="copy" data-sequence="'+sequence+'" title="重新生成文案" aria-label="重新生成文案" onclick="generate(this)">↻</button></div>';const titleNode=cell.querySelector('.copy-title'),titleButton=cell.querySelector('.title-copy-btn');if(titleNode)titleNode.textContent=title;if(titleButton)titleButton.dataset.copy=title;cell.querySelector('.copy-value').textContent=copy;cell.querySelector('.body-copy-btn').dataset.copy=copy}
function setCoverCell(cell,value,sequence){cell.innerHTML='<a href="'+value+'" target="_blank"><img src="'+value+'" alt="封面" loading="lazy"></a><br><button class="copy-btn" type="button" data-action="cover" data-sequence="'+sequence+'" title="重新生成封面" aria-label="重新生成封面" onclick="generate(this)">↻</button>'}
renderPage();
async function generate(button){const action=button.dataset.action,sequence=button.dataset.sequence;const row=button.closest('tr')||document.querySelector('#rows tr[data-sequence="'+sequence+'"]');if(!row)return;const copyButton=row.querySelector('[data-action="copy"]');const copyPending=action==='cover'&&copyButton&&copyButton.classList.contains('generate-btn');if(copyPending){copyButton.disabled=true;copyButton.textContent='生成中…'}button.disabled=true;button.textContent='生成中…';try{const response=await fetch(API_BASE+'/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({action,sequence})});const result=await response.json();if(!response.ok||!result.ok)throw new Error(result.error||'生成失败');const currentRow=document.querySelector('#rows tr[data-sequence="'+sequence+'"]')||row;const currentCopyButton=currentRow.querySelector('[data-action="copy"]');const currentCoverButton=currentRow.querySelector('[data-action="cover"]');if(action==='copy'){const cell=currentCopyButton?currentCopyButton.closest('td'):currentRow.cells[6];if(cell)setCopyCell(cell,result.value,sequence)}else{if(result.copy){const copyCell=currentCopyButton?currentCopyButton.closest('td'):currentRow.cells[6];if(copyCell)setCopyCell(copyCell,result.copy,sequence)}const cell=currentCoverButton?currentCoverButton.closest('td'):currentRow.cells[7];if(cell)setCoverCell(cell,result.value,sequence)}}catch(error){const currentRow=document.querySelector('#rows tr[data-sequence="'+sequence+'"]')||row;const currentButton=currentRow.querySelector('[data-action="'+action+'"]');if(currentButton){currentButton.disabled=false;currentButton.textContent=action==='copy'?'生成文案':'生成封面'}if(copyPending){const currentCopyButton=currentRow.querySelector('[data-action="copy"]');if(currentCopyButton){currentCopyButton.disabled=false;currentCopyButton.textContent='生成文案'}}alert(error.message+'\\n请确认已启动 pdf-catalog serve，并检查 AI 配置。')}};</script></body></html>""", encoding="utf-8")
        # errors 按处理阶段记录；同一 PDF 可能同时在文案、封面等阶段失败。
        # 汇总时按 PDF 路径去重，避免一个 PDF 的多条错误导致成功数变成负数。
        failed_paths = {str(error.get("path", "")) for error in errors if error.get("path")}
        failed_count = len(failed_paths)
        summary = f"发现数: {len(rows)}\n成功数: {max(0, len(rows)-failed_count)}\n失败数: {failed_count}\n耗时秒: {elapsed:.2f}\n"
        detail_text = "\n".join(details or [])
        (settings.output_root/"run.log").write_text((detail_text + "\n\n" if detail_text else "") + summary, encoding="utf-8")
    except PermissionError as exc:
        raise PermissionError(f"输出文件被占用，请关闭 Excel/编辑器后重试: {exc.filename}") from exc

def run(settings: Settings, mode="run", limit=None, no_watermark=False, max_pages=None, dry_run=False, start=1, end=None, ai_start=None, ai_end=None, ai_limit=None, generate_copy_flag=False, generate_cover_flag=False) -> dict[str,int]:
    if not settings.source_root.exists(): raise FileNotFoundError(f"源目录不存在: {settings.source_root}")
    LOG.info("开始扫描 PDF 文件: %s", settings.source_root)
    discovered = discover(settings.source_root)
    LOG.info("文件扫描完成: 共发现 %d 个 PDF", len(discovered))
    lim = limit if limit is not None else settings.processing.get("max_pdfs")
    range_start = max(1, int(start or 1)); end = int(end) if end is not None else None
    files = discovered[range_start - 1:end]
    if lim is not None: files = files[:int(lim)]
    if lim is not None:
        LOG.info("文件筛选完成: 限制数量=%d，实际处理 %d 个 PDF", lim, len(files))
    else:
        LOG.info("文件筛选完成: 未设置数量限制，实际处理 %d 个 PDF", len(files))
    rows=[]; errors=[]; details=[]; run_start=time.time()
    # 重新生成目录时保留已有的 AI 文案和封面，避免仅为刷新 HTML 而丢失已完成内容。
    existing_rows_by_name: dict[str, dict[str, Any]] = {}
    existing_csv = settings.output_root / settings.table.get("csv", "pdf_catalog.csv")
    if existing_csv.is_file():
        try:
            with existing_csv.open(encoding="utf-8-sig", newline="") as f:
                for old_row in csv.DictReader(f):
                    name = str(old_row.get("PDF 文件名称", "") or "").strip()
                    if name:
                        existing_rows_by_name[name] = old_row
        except (OSError, csv.Error) as exc:
            LOG.warning("读取已有目录以保留文案/封面失败，将按空值重新生成: %s", exc)
    start_detail = f"开始处理: 源目录={settings.source_root}，待处理 PDF {len(files)} 个，模式={mode}"
    LOG.info(start_detail)
    details.append(start_detail)
    for idx, pdf in enumerate(files):
        if dry_run:
            detail = f"[{idx + 1}/{len(files)}] 扫描完成: {pdf}"
            LOG.info(detail)
            details.append(detail)
            continue
        if mode in {"scan", "ai"}: title, paths, err = pdf.stem, [], None
        else: title, paths, err = process_pdf(pdf, settings, range_start + idx, not no_watermark)
        if err:
            errors.append({"path":str(pdf),"stage":"process","error":err})
            detail = f"[{idx + 1}/{len(files)}] 失败: {pdf} ({err})"
            LOG.error(detail)
        else:
            detail = f"[{idx + 1}/{len(files)}] 完成: {pdf}，生成图片 {len(paths)} 张"
            LOG.info(detail)
        details.append(detail)
        semester, subject, category = directory_fields(pdf, settings)
        old_row = existing_rows_by_name.get(title, {})
        row={"序号": range_start + idx, "年级":settings.grade,"学期":semester,"科目":subject,"分类":category,"PDF 文件名称":title,
             "生成文案":str(old_row.get("生成文案", "") or ""), "封面图链接":str(old_row.get("封面图链接", "") or "")}
        for i,h in enumerate(HEADERS[8:]): row[h]=paths[i] if i<len(paths) else ""
        ai_index = range_start + idx
        in_ai_range = (ai_start is None or ai_index >= int(ai_start)) and (ai_end is None or ai_index <= int(ai_end))
        if ai_limit is not None and idx >= int(ai_limit): in_ai_range = False
        if in_ai_range and (generate_copy_flag or settings.ai.get("generate_copy")):
            try: row["生成文案"] = generate_copy(settings, pdf.name, category, f"{settings.grade}{subject}")
            except Exception as exc:
                error = f"{type(exc).__name__}: {exc}"
                errors.append({"path":str(pdf),"stage":"copy","error":error})
                LOG.error("[%d/%d] 文案生成失败: %s (%s)", idx + 1, len(files), pdf, error)
                details.append(f"[{idx + 1}/{len(files)}] 文案生成失败: {pdf} ({error})")
        if in_ai_range and (generate_cover_flag or settings.ai.get("generate_cover")):
            try:
                cover_copy = row["生成文案"] or title
                # PDF 已转换出的页面图作为服装/版式参考；有页面图时与配置的主图合并为多图输入。
                cover_url = generate_cover(settings, cover_copy, title, category, f"{settings.grade}{subject}", paths[:5])
                row["封面图链接"] = _download_cover(cover_url, pdf, settings, range_start + idx)
            except Exception as exc:
                error = f"{type(exc).__name__}: {exc}"
                errors.append({"path":str(pdf),"stage":"cover","error":error})
                LOG.error("[%d/%d] 封面生成失败: %s (%s)", idx + 1, len(files), pdf, error)
                details.append(f"[{idx + 1}/{len(files)}] 封面生成失败: {pdf} ({error})")
        rows.append(row)
    if not dry_run:
        LOG.info("开始写入目录文件: %s", settings.output_root)
        write_tables(rows, settings, errors, time.time()-run_start, details)
        LOG.info("目录文件写入完成: %s", settings.output_root)
    # errors 是阶段级明细，统计结果应按 PDF 去重。
    failed_paths = {str(error.get("path", "")) for error in errors if error.get("path")}
    failed_count = len(failed_paths)
    return {"发现数":len(files),"成功数":max(0, len(rows)-failed_count),"失败数":failed_count,"生成图片数":sum(sum(bool(r[h]) for h in HEADERS[8:]) for r in rows)}


def serve(settings: Settings, host: str = "127.0.0.1", port: int = 8765) -> None:
    """启动本地目录服务，为 HTML 中的生成按钮提供安全的本机接口。"""
    output_root = settings.output_root.resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    # 生成过程会读写整份 CSV/HTML；串行化可避免用户同时点击多行时后写入的旧快照覆盖新结果。
    generation_lock = threading.Lock()

    class Handler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(output_root), **kwargs)

        def _json(self, status: int, payload: dict[str, Any]) -> None:
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers(); self.wfile.write(data)

        def do_OPTIONS(self):
            self.send_response(204); self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Headers", "Content-Type"); self.end_headers()

        def do_POST(self):
            if self.path not in {"/api/generate", "/api/status", "/api/open-folder"}:
                self._json(404, {"ok": False, "error": "接口不存在"}); return
            try:
                length = int(self.headers.get("Content-Length", "0")); request = json.loads(self.rfile.read(length))
                with generation_lock:
                    if self.path == "/api/open-folder":
                        sequence = int(request.get("sequence"))
                        if sequence < 1: raise ValueError("参数错误")
                        pdfs = discover(settings.source_root)
                        if sequence > len(pdfs): raise ValueError("找不到对应 PDF")
                        folder = _output_folder(pdfs[sequence - 1], settings, sequence).resolve()
                        if not folder.is_dir(): raise ValueError("文件所在目录不存在")
                        if sys.platform.startswith("win"):
                            os.startfile(str(folder))
                        elif sys.platform == "darwin":
                            subprocess.Popen(["open", str(folder)])
                        else:
                            subprocess.Popen(["xdg-open", str(folder)])
                        self._json(200, {"ok": True}); return
                    if self.path == "/api/status":
                        sequence = int(request.get("sequence")); status = request.get("status")
                        if sequence < 1 or status not in {"已发布", "未发布"}: raise ValueError("参数错误")
                        statuses = _load_publish_status(settings); statuses[str(sequence)] = status
                        _save_publish_status(settings, statuses)
                        _update_status_in_html(settings, sequence, status)
                        self._json(200, {"ok": True, "status": status}); return
                    action, sequence = request.get("action"), int(request.get("sequence"))
                    if action not in {"copy", "cover"} or sequence < 1: raise ValueError("参数错误")
                    pdfs = discover(settings.source_root)
                    if sequence > len(pdfs): raise ValueError("找不到对应 PDF")
                    pdf = pdfs[sequence - 1]
                    csv_path = output_root / settings.table["csv"]
                    with csv_path.open(encoding="utf-8-sig", newline="") as f: rows = list(csv.DictReader(f))
                    row = next((r for r in rows if int(r.get("序号", 0)) == sequence), None)
                    if row is None: raise ValueError("目录中找不到对应行，请先重新生成目录")
                    response_copy = ""
                    if action == "copy":
                        row["生成文案"] = generate_copy(settings, pdf.name, row.get("分类", ""), f"{settings.grade}{row.get('科目', '')}")
                        value = row["生成文案"]
                    else:
                        pages = [row.get(f"图 {i}", "") for i in range(1, 6) if row.get(f"图 {i}", "")]
                        # 封面依赖文案；HTML 直接点击“生成封面”时，先为缺失文案的条目补全文案。
                        copy = str(row.get("生成文案") or "").strip()
                        if not copy:
                            copy = generate_copy(settings, pdf.name, row.get("分类", ""), f"{settings.grade}{row.get('科目', '')}")
                            row["生成文案"] = copy
                        # 页面可能比 CSV 旧：即使文案早已存在，也要随封面结果回传，
                        # 以便前端结束“生成中”状态并展示真实内容。
                        response_copy = copy
                        cover_url = generate_cover(settings, copy, row.get("PDF 文件名称", pdf.stem), row.get("分类", ""), f"{settings.grade}{row.get('科目', '')}", pages)
                        row["封面图链接"] = _download_cover(cover_url, pdf, settings, sequence)
                        value = row["封面图链接"]
                    write_tables(rows, settings, [], 0, [f"HTML 按钮生成{action}: {pdf}"])
                self._json(200, {"ok": True, "value": value, "copy": response_copy})
            except Exception as exc:
                LOG.exception("HTML 生成请求失败")
                # WinError 10013 通常由 Windows 防火墙/代理拦截 Python 访问 Ark 接口，
                # 补充明确提示，避免页面只显示“请确认已启动服务”。
                detail = f"{type(exc).__name__}: {exc}"
                cause = exc
                while getattr(cause, "__cause__", None):
                    cause = cause.__cause__
                if isinstance(exc, urllib.error.URLError) or isinstance(cause, urllib.error.URLError):
                    reason = getattr(cause, "reason", None) or getattr(exc, "reason", None)
                    if getattr(reason, "winerror", None) == 10013 or "10013" in str(reason):
                        detail += "；Windows 阻止了 Python 的网络连接，请在防火墙/安全软件中允许 .venv\\Scripts\\python.exe 访问网络，并检查代理设置"
                self._json(500, {"ok": False, "error": detail})

    server = ThreadingHTTPServer((host, port), Handler)
    LOG.info("目录服务已启动: http://%s:%d/", host, port)
    print(f"目录服务已启动，请打开 {output_root / settings.table.get('html', 'pdf_catalog.html')}")
    try: server.serve_forever()
    except KeyboardInterrupt: pass
    finally: server.server_close()
